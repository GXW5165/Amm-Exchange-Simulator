from __future__ import annotations

"""Excel 多 Sheet 导出模块。

将一次仿真的全部产物（事件记录、摘要指标、用户 PnL、LP 指标、池深度、
图表）导出为一个 .xlsx 文件，每个数据类型占用一个独立的 Sheet。
"""

from dataclasses import asdict
from pathlib import Path
from typing import Any

from src.analytics.pool_depth import compute_pool_depth_curve
from src.simulator.result import SimulationResult


# SimulationArtifacts 延迟导入以避免循环
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from src.application.simulation_runner import SimulationArtifacts


def export_to_excel(
    artifacts: "SimulationArtifacts",
    path: str | Path,
    include_charts: bool = True,
) -> Path:
    """将仿真产物导出为多 Sheet Excel 工作簿。

    Sheet 结构：
        - "Event Records"：所有事件记录行
        - "Summary"：汇总指标
        - "User PnL"：用户盈亏明细
        - "LP Metrics"：LP 做市综合指标
        - "Pool Depth"：池深度曲线数据
        - "Parameters"：仿真配置参数快照
        - "Charts"：嵌入图表图片（当 include_charts=True 且有图时）

    Args:
        artifacts: 仿真产物。
        path: 目标 .xlsx 文件路径。
        include_charts: 是否在工作簿中嵌入图表图片。

    Returns:
        写入后的文件 Path。
    """
    import openpyxl

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # 删除默认空 sheet
    wb.remove(wb.active)

    result = artifacts.result
    summary = result.summary

    # ── Sheet 1: Event Records ──
    _write_events_sheet(wb, result.records)

    # ── Sheet 2: Summary ──
    _write_summary_sheet(wb, summary)

    # ── Sheet 3: User PnL ──
    _write_user_pnl_sheet(wb, summary.user_pnl)

    # ── Sheet 4: LP Metrics ──
    _write_lp_metrics_sheet(wb, result)

    # ── Sheet 5: Pool Depth ──
    _write_pool_depth_sheet(wb, result.pool)

    # ── Sheet 6: Parameters ──
    _write_parameters_sheet(wb, result)

    # ── Sheet 7: Charts (图片) ──
    if include_charts and artifacts.plot_paths:
        artifacts.warnings.extend(_embed_chart_images(wb, artifacts.plot_paths))

    wb.save(output_path)
    return output_path


# ── 内部辅助函数 ──────────────────────────────────────────────────────────


def _auto_column_width(ws) -> None:
    """根据内容自动调整列宽（近似）。"""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted


def _header_style(ws, num_cols: int) -> None:
    """给第一行设置加粗 + 灰底 header 样式。"""
    import openpyxl.styles

    header_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font


def _write_dict_rows(ws, rows: list[dict[str, Any]]) -> None:
    """将字典列表写入 worksheet，首行为列名。"""
    if not rows:
        ws.cell(row=1, column=1, value="(no data)")
        return
    fieldnames = list(rows[0].keys())
    # 写 header
    for col_idx, name in enumerate(fieldnames, 1):
        ws.cell(row=1, column=col_idx, value=name)
    # 写数据行
    for row_idx, row in enumerate(rows, 2):
        for col_idx, name in enumerate(fieldnames, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(name))
    _header_style(ws, len(fieldnames))
    _auto_column_width(ws)


def _write_events_sheet(wb, records: list) -> None:
    """写入事件记录 Sheet。"""
    ws = wb.create_sheet("Event Records")
    rows = [r.to_csv_row() for r in records]
    _write_dict_rows(ws, rows)


def _write_summary_sheet(wb, summary) -> None:
    """写入汇总指标 Sheet（转置为两列：Metric / Value）。"""
    ws = wb.create_sheet("Summary")
    data = asdict(summary)
    # user_pnl 和 lp_annualized_returns 是嵌套结构，单独放在其他 sheet
    data.pop("user_pnl", None)
    data.pop("lp_annualized_returns", None)
    ws.cell(row=1, column=1, value="Metric")
    ws.cell(row=1, column=2, value="Value")
    for row_idx, (key, val) in enumerate(data.items(), 2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=val)
    _header_style(ws, 2)
    _auto_column_width(ws)


def _write_user_pnl_sheet(wb, user_pnl: dict) -> None:
    """写入用户 PnL Sheet。"""
    ws = wb.create_sheet("User PnL")
    rows = [asdict(item) for item in user_pnl.values()]
    _write_dict_rows(ws, rows)


def _write_lp_metrics_sheet(wb, result: SimulationResult) -> None:
    """写入 LP 做市指标 Sheet。

    因为 LP metrics 在 summary 中暂不直接暴露，这里直接调用分析函数复算；
    数据已在 summarize_records 中计算过一次，重复调用的开销可接受。
    """
    from src.analytics.lp_metrics import compute_lp_metrics

    ws = wb.create_sheet("LP Metrics")
    summary = result.summary
    current_price = result.pool.spot_price
    initial_price = result.initial_pool.spot_price

    lp_metrics = compute_lp_metrics(
        records=result.records,
        initial_pool=result.initial_pool,
        current_pool=result.pool,
        initial_users=result.initial_users,
        current_users=result.users,
        price_y_per_x=current_price,
        initial_price_y_per_x=initial_price,
        total_fees_in_y=summary.total_fees_in_y,
    )
    if not lp_metrics:
        ws.cell(row=1, column=1, value="(no LP data)")
        return
    rows = [asdict(m) for m in lp_metrics.values()]
    _write_dict_rows(ws, rows)


def _write_pool_depth_sheet(wb, pool) -> None:
    """写入池深度曲线 Sheet。"""
    ws = wb.create_sheet("Pool Depth")
    depth_points = compute_pool_depth_curve(pool)
    rows = [asdict(dp) for dp in depth_points]
    _write_dict_rows(ws, rows)


def _write_parameters_sheet(wb, result: SimulationResult) -> None:
    """写入仿真关键参数快照 Sheet。"""
    ws = wb.create_sheet("Parameters")
    pool = result.pool
    initial_pool = result.initial_pool
    params = {
        "initial_reserve_x": initial_pool.reserve_x,
        "initial_reserve_y": initial_pool.reserve_y,
        "final_reserve_x": pool.reserve_x,
        "final_reserve_y": pool.reserve_y,
        "fee_rate": pool.fee_rate,
        "initial_total_lp_shares": initial_pool.total_lp_shares,
        "final_total_lp_shares": pool.total_lp_shares,
        "initial_spot_price": initial_pool.spot_price,
        "final_spot_price": pool.spot_price,
        "initial_invariant": initial_pool.invariant,
        "final_invariant": pool.invariant,
        "num_users": len(result.users),
        "num_records": len(result.records),
    }
    ws.cell(row=1, column=1, value="Parameter")
    ws.cell(row=1, column=2, value="Value")
    for row_idx, (key, val) in enumerate(params.items(), 2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=val)
    _header_style(ws, 2)
    _auto_column_width(ws)


def _embed_chart_images(wb, plot_paths: dict[str, Path]) -> list[str]:
    """将图表 PNG 图片嵌入 Excel 的 "Charts" Sheet，并返回失败告警。"""
    import openpyxl.drawing.image

    ws = wb.create_sheet("Charts")
    warnings: list[str] = []
    row = 1
    for name in sorted(plot_paths.keys()):
        path = plot_paths[name]
        if not path.exists():
            continue
        try:
            img = openpyxl.drawing.image.Image(str(path))
            # 缩放到合适尺寸
            img.width = min(img.width, 600)
            img.height = min(img.height, 400)
            ws.add_image(img, f"A{row}")
            ws.cell(row=row, column=1, value=name)
            row += 25  # 每个图约占 25 行
        except Exception as exc:
            warnings.append(f"Excel chart embedding failed for {name}: {exc}")
    return warnings
